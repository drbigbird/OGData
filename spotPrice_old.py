# Download O&G spot price data from EIA

import openpyxl, sys, datetime, requests, os, xlrd
from datetime import datetime

# Create lookup dictionary for converting commodity abbreviations
# (i.e., b - Brent, w - WTI, h - Henry Hub) to file name identifiers
# used on EIA website
abbrDict = {'b': 'pet/hist_xls/RBRTE', 'w': 'pet/hist_xls/RWTC', \
	'h': 'ng/hist_xls/RNGWHHD'}

# Create set containing valid time frequency options
freqOpt = {'d', 'w', 'm', 'a'}

# Define temp xls and final output xls names
tempfn = 'tempPriceData.xls'
exportfn = 'PriceData.xlsx'

def setOptions(args):
	"""Sets the options for what price to download
	with the arguments:
		comType - Commodity type
		freq - Frequency for pricing (mm/dd/yyyy)
		startDate - Starting date"""

	comType = ''
	freq = ''
	startDate = None
	dateInput = ''

	# Get commodity type
	if len(args) > 0 and args[0] in abbrDict:
		comType = args[0]
	else:
		while comType not in abbrDict:
			comType = input('Commodity (b - Brent/w - WTI/h - Henry Hub): ')

	# Get frequency
	if len(args) > 1 and args[1] in freqOpt:
		freq = args[1]
	else:
		while freq not in freqOpt:
			freq = input('Frequency (d - day/weekly - w/m - monthly/a - annual): ')

	# Get start date
	try:
		startDate = datetime.strptime(args[2],'%m/%d/%Y')
	except:
		pass
	
	while startDate == None:

		dateInput = input('Starting date (mm/dd/yyyy): ')
		
		try:
			startDate = datetime.strptime(dateInput,'%m/%d/%Y')
		except:
			if dateInput == '':
				startDate = datetime(1900,1,1)

	return comType, freq, startDate

def createOutput(tempfn, startDate):
	"""Selects relevant price data and creates final output file
	with arguments:
		tempfn - Filename of the temporary downloaded file
		startDate - Start date for the price data"""

	wb = xlrd.open_workbook(tempfn)
	dm = wb.datemode
	sheet = wb.sheet_by_index(1)
	pdict = {}

	for i in range(3,sheet.nrows):
		
		tt = xlrd.xldate_as_tuple(sheet.cell(i,0).value,dm)
		dateData = datetime(tt[0],tt[1],tt[2])
		priceData = sheet.cell(i,1).value

		if dateData >= startDate:
			pdict[dateData] = priceData

	clabels = [sheet.cell(2,0).value, sheet.cell(2,1).value]

	return pdict, clabels

def exportData(exportfn, pdict, clabels):
	
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = 'Historical prices'

	# Create lists from the dictionary data
	dates = list(pdict.keys())
	prices = list(pdict.values())

	# Write the column headers
	sheet.cell(2,2,clabels[0])
	sheet.cell(2,3,clabels[1])

	# Populate price data
	for i in range(0,len(pdict)):

		sheet.cell(i+3,2,dates[i].date())
		sheet.cell(i+3,3,prices[i])

	wb.save(exportfn)


# Get download options and construct URL string
comType, freq, startDate = setOptions(sys.argv[1:])
url = 'https://www.eia.gov/dnav/' + abbrDict[comType] + freq + '.xls'

# Download the data file from EIA
res = requests.get(url)
with open(tempfn, 'wb') as output:
	output.write(res.content)
output.close()

pdict, clabels = createOutput('temp.xls',datetime(1888,3,15))
exportData(exportfn, pdict, clabels)

# Remove temporary file
os.remove(tempfn)
